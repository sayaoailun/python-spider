import openpyxl
import random
import time
import queue
import threading
import csv
import json
import re
import bs4
import requests
import sys
import io
import pprint
import logging
sys.stdout = io.TextIOWrapper(
    sys.stdout.buffer, encoding='utf8')  # 改变标准输出的默认编码

fmt = '%(asctime)s %(pathname)s[line:%(lineno)d] %(module)s.%(funcName)s [%(processName)s/%(threadName)s] %(levelname)s : %(message)s'
# logging.basicConfig(handlers = [logging.FileHandler('log.txt', mode = 'w', encoding = 'utf-8')], level = logging.INFO, format = '%(asctime)s - %(levelname)s - %(message)s')
logging.basicConfig(handlers=[logging.StreamHandler(sys.stdout), logging.FileHandler(
    'log.txt', mode='w', encoding='utf-8')], level=logging.INFO, format=fmt)


threadPoolSize = 1
threadPool = queue.Queue()
bookQueue = queue.Queue()
threads = []
sleepTime = 10

count = 0
cookie = []
_lock = threading.RLock()

def getSleepTime() :
    return random.randint(1,10)

def getCookie():
    user_agent_list = [
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/22.0.1207.1 Safari/537.1",
        "Mozilla/5.0 (X11; CrOS i686 2268.111.0) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.57 Safari/536.11",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1092.0 Safari/536.6",
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1090.0 Safari/536.6",
        "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/19.77.34.5 Safari/537.1",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.9 Safari/536.5",
        "Mozilla/5.0 (Windows NT 6.0) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.36 Safari/536.5",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3",
        "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_8_0) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3",
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1062.0 Safari/536.3",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1062.0 Safari/536.3",
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3",
        "Mozilla/5.0 (Windows NT 6.1) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3",
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.0 Safari/536.3",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/535.24 (KHTML, like Gecko) Chrome/19.0.1055.1 Safari/535.24",
        "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/535.24 (KHTML, like Gecko) Chrome/19.0.1055.1 Safari/535.24"
    ]
    url = 'https://www.amazon.cn'
    _lock.acquire()
    cookie.clear()
    while len(cookie) < threadPoolSize:
        UA = random.choice(user_agent_list)
        headers = {'User-Agent': UA}
        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            logging.info(response.headers)
            if not response.headers.get('Set-Cookie') == None:
                cookie.append(response.headers['Set-Cookie'])
            time.sleep(getSleepTime())
        except Exception as e:
            logging.exception(e)
            time.sleep(getSleepTime())
    _lock.release()
    logging.info(cookie)


def getThread():
    while True:
        if threadPool.qsize() < threadPoolSize:
            break
        else:
            time.sleep(getSleepTime())


class Spider:
    def __init__(self, url, filename):
        self.url = url
        self.amazonUrl = 'https://www.amazon.cn'
        self.filename = filename
        self.books = []
        self.user_agent_list = [
            "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/22.0.1207.1 Safari/537.1",
            "Mozilla/5.0 (X11; CrOS i686 2268.111.0) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.57 Safari/536.11",
            "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1092.0 Safari/536.6",
            "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1090.0 Safari/536.6",
            "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/19.77.34.5 Safari/537.1",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.9 Safari/536.5",
            "Mozilla/5.0 (Windows NT 6.0) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.36 Safari/536.5",
            "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3",
            "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_8_0) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3",
            "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1062.0 Safari/536.3",
            "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1062.0 Safari/536.3",
            "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3",
            "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3",
            "Mozilla/5.0 (Windows NT 6.1) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3",
            "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.0 Safari/536.3",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/535.24 (KHTML, like Gecko) Chrome/19.0.1055.1 Safari/535.24",
            "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/535.24 (KHTML, like Gecko) Chrome/19.0.1055.1 Safari/535.24"
        ]

    def save(self):
        self.bookList()
        file = open(self.filename, 'w', newline='')
        writer = csv.writer(file)
        writer.writerow(['排名', '书名', '价格', '作者', '出版社',
                         'ASIN', '文件大小', '纸书页数', '语种', '品牌'])
        for book in self.books:
            writer.writerow([book['num'], book['name'], book['price'], book['authors'], book['press'],
                             book['asin'], book['size'], book['page'], book['language'], book['brand']])
        file.close()

    def appendBook(self):
        self.bookList()
        file = open(self.filename, 'a', newline='')
        writer = csv.writer(file)
        for book in self.books:
            writer.writerow([book['num'], book['name'], book['price'], book['authors'], book['press'],
                             book['asin'], book['size'], book['page'], book['language'], book['brand']])
        file.close()

    def fulfillBook(self, book):
        if book.get('brand') == None:
            book['brand'] = '-'
        if book.get('size') == None:
            book['size'] = '-'
        if book.get('page') == None:
            book['page'] = '-'
        if book.get('asin') == None:
            book['asin'] = '-'
        if book.get('language') == None:
            book['language'] = '-'
        if book.get('press') == None:
            book['press'] = '-'
        if book.get('url') == None:
            book['url'] = '-'
        if book.get('name') == None:
            book['name'] = '-'
        if book.get('price') == None:
            book['price'] = '-'
        if book.get('num') == None:
            book['num'] = '-'

    def tagList(self):
        UA = random.choice(self.user_agent_list)
        headers = {'User-Agent': UA}
        response = requests.get(self.url, headers=headers)
        try:
            response.raise_for_status()
        except Exception as e:
            if str(e).startswith('404'):
                return []
        response.encoding = 'utf-8'
        soup = bs4.BeautifulSoup(response.text, 'html.parser')
        taglist = soup.select(
            'div[class="a-section a-spacing-none aok-relative"]')
        return taglist

    def bookInfo(self, bookTag):
        book = {}
        soup = bs4.BeautifulSoup(str(bookTag), 'html.parser')
        try:
            book['url'] = self.amazonUrl + soup.select(
                'span[class="aok-inline-block zg-item"] > a[class="a-link-normal"]')[0].get('href')
            book['name'] = soup.select('span > div > img')[0].get("alt").replace('・', '·').replace(
                '•', '·').replace('\u2219', '·').replace('\u25aa', '·').replace('®', '').replace('ë', ' ')
            if len(soup.select('span[class="p13n-sc-price"]')) > 0:
                book['price'] = soup.select(
                    'span[class="p13n-sc-price"]')[0].getText().replace('￥', '')
            book['num'] = soup.select(
                'span[class="zg-badge-text"]')[0].getText().replace('#', '')
        except Exception as e:
            logging.info(soup)
            # raise e
            fulfillBook(book)

        # 获取书籍详情
        notfind = True
        soup = None
        while notfind:
            UA = random.choice(self.user_agent_list)
            headers = {'User-Agent': UA}
            headers['Cookie'] = random.choice(cookie)
            try:
                response = requests.get(book['url'], headers=headers)
                response.raise_for_status()
            except Exception as e:
                logging.exception(e)
                time.sleep(getSleepTime())
                notfind = True
                continue
            response.encoding = 'utf-8'
            soup = bs4.BeautifulSoup(response.text, 'html.parser')
            if len(soup.select('div[class="a-section a-spacing-micro bylineHidden feature"]')) == 0:
                getCookie()
                time.sleep(getSleepTime())
                _lock.acquire()
                global count
                count += 1
                logging.info('oops!%s' % count)
                _lock.release()
                notfind = True
                continue
            else:
                notfind = False
        book['authors'] = ''
        for authorTag in soup.select('span[class="author notFaded"]'):
            book['authors'] += authorTag.getText().replace('\n', '').replace('・', '·').replace(
                '•', '·').replace('\u2219', '·').replace('\u25aa', '·').replace('®', '').replace('ë', ' ')
        spanTags = soup.select('div[id="detailBullets_feature_div"] > ul > li > span')
        for spanTag in spanTags:
          spans = spanTag.select('span > span')
          if len(spans) == 2:
            if '文件大小' in spans[0].getText():
              book['size'] = spans[1].getText()
              continue
            if '纸书页数' in spans[0].getText():
              book['page'] = spans[1].getText()
              continue
            if '出版社' in spans[0].getText():
              book['press'] = spans[1].getText()
              continue
            if '语种' in spans[0].getText():
              book['language'] = spans[1].getText()
              continue
            if 'ASIN' in spans[0].getText():
              book['asin'] = spans[1].getText()
              continue
            if '品牌' in spans[0].getText():
              book['brand'] = spans[1].getText().replace('・', '·').replace('•', '·').replace(
                    '\u2219', '·').replace('\u25aa', '·').replace('®', '').replace('ë', ' ')
              continue
        self.fulfillBook(book)
        bookQueue.put(book)
        threadPool.get()
        logging.info(book)
        return book

    def bookList(self):
        taglist = self.tagList()
        for tag in taglist:
            self.books.append(self.bookInfo(tag))
        return self.books

    def bookQueue(self):
        taglist = self.tagList()
        for tag in taglist:
            getThread()
            thread = threading.Thread(target=self.bookInfo, args=[tag])
            thread.start()
            threadPool.put(thread)
            threads.append(thread)
        return taglist


def write(sheetAndUrl, file):
    wb = openpyxl.Workbook()
    for sheet in sorted(sheetAndUrl):
        # 付费
        urlTemplate = sheetAndUrl[sheet]
        for x in range(1, 3):
            spider = Spider(urlTemplate+'?pg=%s' % x, sheet)
            if len(spider.bookQueue()) == 0:
                break
        for x in threads:
            x.join()
        bookDict = {}
        for x in range(bookQueue.qsize()):
            book = bookQueue.get()
            bookDict[book['num']] = book
        ws = wb.create_sheet(sheet+'付费排行榜')
        title = ['排名', '书名', '价格', '作者', '出版社',
                 'ASIN', '文件大小', '纸书页数', '语种', '品牌']
        for x in range(1, len(title) + 1):
            c = ws.cell(row=1, column=x)
            c.value = title[x-1]
        for x in range(2, len(bookDict) + 2):
            book = bookDict['%s' % (x - 1)]
            content = [book['num'], book['name'], book['price'], book['authors'], book['press'],
                       book['asin'], book['size'], book['page'], book['language'], book['brand']]
            for i in range(len(content)):
                c = ws.cell(row=x, column=i+1)
                c.value = content[i]
                if i == 1:
                    c.style = 'Hyperlink'
                    c.hyperlink = book['url']
        # 免费
        urlTemplate = sheetAndUrl[sheet]
        for x in range(1, 3):
            spider = Spider(urlTemplate+'?tf=1&pg=%s' % x, sheet)
            if len(spider.bookQueue()) == 0:
                break
        for x in threads:
            x.join()
        bookDict = {}
        for x in range(bookQueue.qsize()):
            book = bookQueue.get()
            bookDict[book['num']] = book
        ws = wb.create_sheet(sheet+'免费排行榜')
        title = ['排名', '书名', '价格', '作者', '出版社',
                 'ASIN', '文件大小', '纸书页数', '语种', '品牌']
        for x in range(1, len(title) + 1):
            c = ws.cell(row=1, column=x)
            c.value = title[x-1]
        for x in range(2, len(bookDict) + 2):
            book = bookDict['%s' % (x - 1)]
            content = [book['num'], book['name'], book['price'], book['authors'], book['press'],
                       book['asin'], book['size'], book['page'], book['language'], book['brand']]
            for i in range(len(content)):
                c = ws.cell(row=x, column=i+1)
                c.value = content[i]
                if i == 1:
                    c.style = 'Hyperlink'
                    c.hyperlink = book['url']
    del wb['Sheet']
    wb.save(file)


if __name__ == '__main__':
    getCookie()
    sheetAndUrl = {}
    sheetAndUrl['电子书'] = 'https://www.amazon.cn/gp/bestsellers/digital-text/116169071'
    sheetAndUrl['小说'] = 'https://www.amazon.cn/gp/bestsellers/digital-text/144154071'
    sheetAndUrl['文学'] = 'https://www.amazon.cn/gp/bestsellers/digital-text/144180071'
    sheetAndUrl['英语'] = 'https://www.amazon.cn/gp/bestsellers/digital-text/143324071'
    sheetAndUrl['原版'] = 'https://www.amazon.cn/gp/bestsellers/digital-text/116170071'
    write(sheetAndUrl, 'amazon-%s.xlsx' %
          time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime(time.time())))
