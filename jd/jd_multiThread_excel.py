import openpyxl
import random
import time
import queue
import threading
import csv
import json
import re
import bs4
import requests
import sys
import io
import pprint
import logging
sys.stdout = io.TextIOWrapper(
    sys.stdout.buffer, encoding='utf8')  # 改变标准输出的默认编码

fmt = '%(asctime)s %(pathname)s[line:%(lineno)d] %(module)s.%(funcName)s [%(processName)s/%(threadName)s] %(levelname)s : %(message)s'
# logging.basicConfig(handlers = [logging.FileHandler('log.txt', mode = 'w', encoding = 'utf-8')], level = logging.INFO, format = '%(asctime)s - %(levelname)s - %(message)s')
logging.basicConfig(handlers=[logging.StreamHandler(sys.stdout), logging.FileHandler(
    'log.txt', mode='w', encoding='utf-8')], level=logging.INFO, format=fmt)


threadPoolSize = 10
threadPool = queue.Queue()
bookQueue = queue.Queue()
threads = []


def getThread():
    while True:
        if threadPool.qsize() < threadPoolSize:
            break
        else:
            time.sleep(3)


class Spider:
    def __init__(self, url, filename):
        self.url = url
        self.filename = filename
        self.books = []
        self.user_agent_list = [
            "PostmanRuntime/7.26.5"
        ]

    def save(self):
        self.bookList()
        file = open(self.filename, 'w', newline='')
        writer = csv.writer(file)
        writer.writerow(['排名', '书名', '价格', '定价', '作者', '出版社', 'ID',
                        'ISBN', '分类', '版次', '开本', '出版时间', '页数', '正文语种'])
        for book in self.books:
            writer.writerow([book['num'], book['name'], book['price'], book['m_price'], book['authors'], book['press'], book['id'],
                            book['isbn'], book['class'], book['version'], book['size'], book['time'], book['page'], book['language']])
        file.close()

    def appendBook(self):
        self.bookList()
        file = open(self.filename, 'a', newline='')
        writer = csv.writer(file)
        for book in self.books:
            writer.writerow([book['num'], book['name'], book['price'], book['m_price'], book['authors'], book['press'], book['id'],
                            book['isbn'], book['class'], book['version'], book['size'], book['time'], book['page'], book['language']])
        file.close()

    def fulfillBook(self, book):
        if book.get('version') == None:
            book['version'] = '-'
        if book.get('size') == None:
            book['size'] = '-'
        if book.get('page') == None:
            book['page'] = '-'
        if book.get('time') == None:
            book['time'] = '-'
        if book.get('language') == None:
            book['language'] = '-'
        if book.get('isbn') == None:
            book['isbn'] = '-'
        if book.get('press') == None:
            book['press'] = '-'

    def tagList(self):
        UA = random.choice(self.user_agent_list)
        headers = {'User-Agent': UA}
        response = requests.get(self.url, headers=headers)
        response.raise_for_status()
        soup = bs4.BeautifulSoup(response.content, 'html.parser')
        taglist = soup.select('ul[class="clearfix"] > li[clstag]')
        return taglist

    def bookInfo(self, bookTag):
        book = {}
        soup = bs4.BeautifulSoup(str(bookTag), 'html.parser')
        book['id'] = soup.select('em[data-price-id]')[0].get('data-price-id')
        book['num'] = soup.select('div[class="p-num"]')[0].getText()
        book['name'] = soup.select(
            'a[class="p-name"]')[0].get('title').replace('\u3000', ' ')
        detail = soup.select('div[class="p-detail"] > dl')
        authors = []
        for author in bs4.BeautifulSoup(str(detail[0]), 'html.parser').select('dd > a'):
            authors.append(author.get('title').replace('\u3000', ' '))
        book['authors'] = authors
        book['press'] = bs4.BeautifulSoup(
            str(detail[1]), 'html.parser').select('dd > a')[0].get('title')

        # 获取 book price
        # UA = random.choice(self.user_agent_list)
        # headers = {'User-Agent': UA}
        # priceUrl = 'http://p.3.cn/prices/get?skuid=%s' % book['id']
        while True:
            try:
                UA = random.choice(self.user_agent_list)
                headers = {'User-Agent': UA}
                priceUrl = 'http://p.3.cn/prices/get?skuid=%s' % book['id']
                priceResponse = requests.get(priceUrl, headers=headers)
                break;
            except Exception as e:
                logging.info(e)
                time.sleep(30)
        # priceResponse = requests.get(priceUrl, headers = headers)
        priceResponse.raise_for_status()
        book['price'] = json.loads(priceResponse.text)[0].get('p')
        book['m_price'] = json.loads(priceResponse.text)[0].get('m')

        # 获取 book ISBN等信息
        while True:
            UA = random.choice(self.user_agent_list)
            headers = {'User-Agent': UA}
            bookUrl = 'http://item.jd.com/%s.html' % book['id']
            book['url'] = bookUrl
            bookResponse = requests.get(bookUrl, headers=headers)
            try:
                bookResponse.raise_for_status()
            except Exception as e:
                logging.info(e)
                time.sleep(30)
                continue
            # bookResponse.raise_for_status()
            # bookResponse.encoding = 'gbk'
            if len(bs4.BeautifulSoup(bookResponse.text, 'html.parser').select('ul[id="parameter2"] > li')) > 0:
                break
            time.sleep(30)
        # UA = random.choice(self.user_agent_list)
        # headers = {'User-Agent': UA}
        # bookUrl = 'http://item.jd.com/%s.html' % book['id']
        # book['url'] = bookUrl
        # bookResponse = requests.get(bookUrl, headers = headers)
        # bookResponse.raise_for_status()
        # bookResponse.encoding = 'gbk'
        # if len(bs4.BeautifulSoup(bookResponse.text, 'html.parser').select('ul[id="parameter2"] > li')) > 0:
        #   break
        for x in bs4.BeautifulSoup(bookResponse.text, 'html.parser').select('ul[id="parameter2"] > li'):
            if x.getText().find('ISBN') != -1:
                book['isbn'] = x.get('title')
                continue
            if x.getText().find('版次') != -1:
                book['version'] = x.get('title')
                continue
            if x.getText().find('开本') != -1:
                book['size'] = x.get('title')
                continue
            if x.getText().find('页数') != -1:
                book['page'] = x.get('title')
                continue
            if x.getText().find('出版时间') != -1:
                book['time'] = x.get('title')
                continue
            if x.getText().find('正文语种') != -1:
                book['language'] = x.get('title')
                continue
        self.fulfillBook(book)
        # 获取分类
        clazz = '图书'
        for x in bs4.BeautifulSoup(bookResponse.text, 'html.parser').select('div[class="crumb fl clearfix"] > div[class="item"] > a'):
            clazz += ('>' + x.getText())
        book['class'] = clazz
        bookQueue.put(book)
        threadPool.get()
        logging.info(book)
        return book

    def bookList(self):
        taglist = self.tagList()
        for tag in taglist:
            self.books.append(self.bookInfo(tag))
        return self.books

    def bookQueue(self):
        taglist = self.tagList()
        for tag in taglist:
            getThread()
            thread = threading.Thread(target = self.bookInfo, args = [tag])
            thread.start()
            threadPool.put(thread)
            threads.append(thread)

def write(sheetAndUrl, file):
    wb = openpyxl.Workbook()
    for sheet in sorted(sheetAndUrl):
        urlTemplate = sheetAndUrl[sheet]
        for x in range(1,6):
            spider = Spider(urlTemplate % x, sheet)
            spider.bookQueue()
        for x in threads:
            x.join()
        bookDict = {}
        for x in range(bookQueue.qsize()):
            book = bookQueue.get()
            bookDict[book['num']] = book
        ws = wb.create_sheet(sheet)
        title = ['排名', '书名', '价格', '定价', '作者', '出版社', 'ID', 'ISBN', '分类', '版次', '开本', '出版时间', '页数', '正文语种']
        for x in range(1,len(title) + 1):
            c = ws.cell(row=1,column=x)
            c.value = title[x-1]
        for x in range(2,len(bookDict) + 2):
            if x - 1 <= 9:
                book = bookDict[' 0%s ' % (x - 1)]
            else:
                book = bookDict[' %s ' % (x - 1)]
            try:
                content = [book['num'], book['name'], book['price'], book['m_price'], str(book['authors']), book['press'], book['id'], book['isbn'], book['class'], book['version'], book['size'], book['time'], book['page'], book['language']]
            except Exception as e:
                logging.info(book)
                raise e
            # content = [book['num'], book['name'], book['price'], book['m_price'], str(book['authors']), book['press'], book['id'], book['isbn'], book['class'], book['version'], book['size'], book['time'], book['page'], book['language']]
            for i in range(len(content)):
                c = ws.cell(row=x,column=i+1)
                c.value = content[i]
                if i == 1:
                    c.style = 'Hyperlink'
                    c.hyperlink = book['url']
    del wb['Sheet']
    wb.save(file)


if __name__ == '__main__':
    sheetAndUrl = {}
    sheetAndUrl['jd图书销量榜'] = 'https://book.jd.com/booktop/0-0-0.html?category=1713-0-0-0-10002-%s#comfort'
    sheetAndUrl['jd小说文学图书销量榜'] = 'https://book.jd.com/booktop/0-0-0.html?category=20002-0-0-0-10002-%s#comfort'
    sheetAndUrl['jd原版图书销量榜'] = 'https://book.jd.com/booktop/0-0-0.html?category=20008-0-0-0-10002-%s#comfort'
    write(sheetAndUrl, 'jd-%s.xlsx' % time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime(time.time())))
