import openpyxl
import re
import sys
import io
import logging
import os
import mysql.connector
sys.stdout = io.TextIOWrapper(
    sys.stdout.buffer, encoding='utf8')  # 改变标准输出的默认编码

fmt = '%(asctime)s %(pathname)s[line:%(lineno)d] %(module)s.%(funcName)s [%(processName)s/%(threadName)s] %(levelname)s : %(message)s'
logging.basicConfig(handlers=[logging.StreamHandler(sys.stdout), logging.FileHandler(
    'log.txt', mode='w', encoding='utf-8')], level=logging.INFO, format=fmt)


def getFileList(dirPath):
    files = os.listdir(dirPath)
    fileList = []
    for file in files:
        filePath = '%s/%s' % (dirPath, file)
        if not os.path.isdir(filePath) and re.match(r'([a-z]+)-(.+)_(.+).xlsx', file):
            matchObj = re.match(r'(.+?)-(.+)_(.+).xlsx', file)
            fileDict = {}
            fileDict['filePath'] = filePath
            fileDict['platform'] = matchObj.group(1)
            fileDict['date'] = '%s_%s' % (matchObj.group(2), matchObj.group(3))
            fileList.append(fileDict)
    return fileList


def saveToDb(dirPath):
    mysql_config = {
        'user': 'root',
        'password': 'mysql',
        'host': '10.126.2.56',
        'port': 3306,
        'database': 'books'
    }

    cnx = mysql.connector.connect(**mysql_config)
    cursor = cnx.cursor()

    amazonColumn = {
        '排名': 1,
        '书名': 2,
        '价格': 3,
        '作者': 4,
        '出版社': 5,
        'ASIN': 6,
        '文件大小': 7,
        '纸书页数': 8,
        '语种': 9,
        '品牌': 10
    }

    jdColumn = {
        '排名': 1,
        '书名': 2,
        '价格': 3,
        '定价': 4,
        '作者': 5,
        '出版社': 6,
        'ID': 7,
        'ISBN': 8,
        '分类': 9,
        '版次': 10,
        '开本': 11,
        '出版时间': 12,
        '页数': 13,
        '正文语种': 14
    }

    dangdangColumn = {
        '排名': 1,
        '书名': 2,
        '价格': 3,
        '定价': 4,
        '作者': 5,
        '出版社': 6,
        'ID': 7,
        'ISBN': 8,
        '分类': 9,
        '版次': 10,
        '开本': 11,
        '出版时间': 12,
        '页数': 13,
        '正文语种': 14
    }

    fileList = getFileList(dirPath)
    for file in fileList:
        if 'amazon' == file['platform']:
            wb = openpyxl.load_workbook(file['filePath'])
            insert = ('insert into amazon '
                      '(rank, book_name, price, author, press, asin, file_size, page, language, brand, link, rank_list, date) '
                      'values (%(rank)s, %(book_name)s, %(price)s, %(author)s, %(press)s, %(asin)s, %(file_size)s, %(page)s, %(language)s, %(brand)s, %(link)s, %(rank_list)s, %(date)s)')
            for sheet in wb.sheetnames:
                for row in range(2, wb[sheet].max_row+1):
                    book = {}
                    book['rank_list'] = sheet
                    book['date'] = file['date']
                    book['rank'] = wb[sheet].cell(
                        row=row, column=amazonColumn['排名']).value
                    book['book_name'] = wb[sheet].cell(
                        row=row, column=amazonColumn['书名']).value
                    book['link'] = wb[sheet].cell(
                        row=row, column=amazonColumn['书名']).hyperlink.target
                    book['price'] = wb[sheet].cell(
                        row=row, column=amazonColumn['价格']).value.replace(",", "")
                    book['author'] = wb[sheet].cell(
                        row=row, column=amazonColumn['作者']).value
                    book['press'] = wb[sheet].cell(
                        row=row, column=amazonColumn['出版社']).value
                    book['asin'] = wb[sheet].cell(
                        row=row, column=amazonColumn['ASIN']).value
                    book['file_size'] = wb[sheet].cell(
                        row=row, column=amazonColumn['文件大小']).value
                    book['page'] = wb[sheet].cell(
                        row=row, column=amazonColumn['纸书页数']).value.replace("页", "")
                    book['language'] = wb[sheet].cell(
                        row=row, column=amazonColumn['语种']).value
                    book['brand'] = wb[sheet].cell(
                        row=row, column=amazonColumn['品牌']).value
                    logging.info(book)
                    if '-' == book['page']:
                        book['page'] = 0
                    if '-' == book['price']:
                        book['price'] = 0
                    cursor.execute(insert, book)
            wb.close()
        if 'jd' == file['platform']:
            wb = openpyxl.load_workbook(file['filePath'])
            insert = ('insert into jd '
                      '(rank, book_name, item_id, price_real, price_mark, author, press, isbn, book_size, page, language, link, class, publish_time, rank_list, version, date) '
                      'values (%(rank)s, %(book_name)s, %(item_id)s, %(price_real)s, %(price_mark)s, %(author)s, %(press)s, %(isbn)s, %(book_size)s, %(page)s, %(language)s, %(link)s, %(class)s, %(publish_time)s, %(rank_list)s, %(version)s, %(date)s)')
            for sheet in wb.sheetnames:
                for row in range(2, wb[sheet].max_row+1):
                    book = {}
                    book['rank_list'] = sheet
                    book['date'] = file['date']
                    book['rank'] = wb[sheet].cell(
                        row=row, column=jdColumn['排名']).value
                    book['book_name'] = wb[sheet].cell(
                        row=row, column=jdColumn['书名']).value
                    book['link'] = wb[sheet].cell(
                        row=row, column=jdColumn['书名']).hyperlink.target
                    book['item_id'] = wb[sheet].cell(
                        row=row, column=jdColumn['ID']).value
                    book['price_real'] = wb[sheet].cell(
                        row=row, column=jdColumn['价格']).value.replace(",", "")
                    book['price_mark'] = wb[sheet].cell(
                        row=row, column=jdColumn['定价']).value.replace(",", "")
                    book['author'] = wb[sheet].cell(
                        row=row, column=jdColumn['作者']).value
                    book['press'] = wb[sheet].cell(
                        row=row, column=jdColumn['出版社']).value
                    book['isbn'] = wb[sheet].cell(
                        row=row, column=jdColumn['ISBN']).value
                    book['book_size'] = wb[sheet].cell(
                        row=row, column=jdColumn['开本']).value
                    book['page'] = wb[sheet].cell(
                        row=row, column=jdColumn['页数']).value.replace("页", "")
                    book['language'] = wb[sheet].cell(
                        row=row, column=jdColumn['正文语种']).value
                    book['class'] = wb[sheet].cell(
                        row=row, column=jdColumn['分类']).value
                    book['publish_time'] = wb[sheet].cell(
                        row=row, column=jdColumn['出版时间']).value
                    book['version'] = wb[sheet].cell(
                        row=row, column=jdColumn['版次']).value
                    logging.info(book)
                    if '-' == book['page']:
                        book['page'] = '0'
                    if '×' in book['page']:
                        matchObj = re.match(r'([\d]+)×([\d]+)', book['page'])
                        book['page'] = str(int(matchObj.group(1)) * int(matchObj.group(2)))
                    if '页' in book['page']:
                        matchObj = re.match(r'([\d]+)页', book['page'])
                        book['page'] = matchObj.group(1)
                    if 'press' not in book:
                        book['press'] = '-'
                    if '-' == book['price_mark']:
                        book['price_mark'] = 0
                    if '-' == book['price_real']:
                        book['price_real'] = 0
                    cursor.execute(insert, book)
            wb.close()
        if 'dangdang' == file['platform']:
            wb = openpyxl.load_workbook(file['filePath'])
            insert = ('insert into dangdang '
                      '(rank, book_name, item_id, price_real, price_mark, author, press, isbn, book_size, link, class, publish_time, rank_list, date) '
                      'values (%(rank)s, %(book_name)s, %(item_id)s, %(price_real)s, %(price_mark)s, %(author)s, %(press)s, %(isbn)s, %(book_size)s, %(link)s, %(class)s, %(publish_time)s, %(rank_list)s, %(date)s)')
            for sheet in wb.sheetnames:
                for row in range(2, wb[sheet].max_row+1):
                    book = {}
                    book['rank_list'] = sheet
                    book['date'] = file['date']
                    book['rank'] = wb[sheet].cell(
                        row=row, column=dangdangColumn['排名']).value
                    book['book_name'] = wb[sheet].cell(
                        row=row, column=dangdangColumn['书名']).value
                    book['link'] = wb[sheet].cell(
                        row=row, column=dangdangColumn['书名']).hyperlink.target
                    book['item_id'] = wb[sheet].cell(
                        row=row, column=dangdangColumn['ID']).value
                    book['price_real'] = wb[sheet].cell(
                        row=row, column=dangdangColumn['价格']).value.replace(",", "")
                    book['price_mark'] = wb[sheet].cell(
                        row=row, column=dangdangColumn['定价']).value.replace(",", "")
                    book['author'] = wb[sheet].cell(
                        row=row, column=dangdangColumn['作者']).value
                    book['press'] = wb[sheet].cell(
                        row=row, column=dangdangColumn['出版社']).value
                    book['isbn'] = wb[sheet].cell(
                        row=row, column=dangdangColumn['ISBN']).value
                    book['book_size'] = wb[sheet].cell(
                        row=row, column=dangdangColumn['开本']).value
                    book['class'] = wb[sheet].cell(
                        row=row, column=dangdangColumn['分类']).value
                    book['publish_time'] = wb[sheet].cell(
                        row=row, column=dangdangColumn['出版时间']).value
                    if '-' == book['price_mark']:
                        book['price_mark'] = 0
                    if '-' == book['price_real']:
                        book['price_real'] = 0
                    logging.info(book)
                    cursor.execute(insert, book)
            wb.close()


    # Make sure data is committed to the database
    cnx.commit()

    cursor.close()
    cnx.close()


if __name__ == '__main__':
    saveToDb('.')
